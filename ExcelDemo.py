import openpyxl
book = openpyxl.load_workbook("C:\\Users\\voona\\OneDrive\\Desktop\\Practice\\SeleniumImp\\PythonExcel.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
Dict = {}
print(cell.value)

# Write value to cell in Excel
sheet.cell(row=2, column=2).value = "Vikash"
print(sheet.cell(row=2, column=2).value)

print(sheet.max_row, sheet.max_column)
# print(sheet["A4"].value)

for i in range(1, sheet.max_row+1):
    if sheet.cell(row=i, column=1).value == "testcase3":
        for j in range(2, sheet.max_column+1):
            c = sheet.cell(row=i, column=j)
            Dict[sheet.cell(row=1, column=j).value] = c.value
            # print(c.value, end=" ")
        # print()
print(Dict)